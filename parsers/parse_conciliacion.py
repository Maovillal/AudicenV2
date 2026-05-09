"""
Parser para el archivo de conciliación de envase (.xlsx).

El archivo tiene 500+ pestañas. Por convención del supervisor, hay UNA pestaña
por día con nombre dd.mm.yyyy (ej. "08.05.2026"). Pueden existir pestañas
extra al final como "Conteo 1", "Hoja2", etc., pero las ignoramos — el parser
busca específicamente la pestaña que coincide con la fecha del correo.

Cada pestaña tiene dos recuadros pequeños con los totales que importan para
la auditoría:

Recuadro T1:
    Fisico       <valor>     ← se sobrescribe entre T1 inicio (mañana) y
    Fin SAP      <valor>       T1 cierre (tarde, después de 12:00). El subject
    Diferencia   <valor>       del correo distingue cuándo se capturó.

Recuadro T2:
    Inicio SAP        <valor>
    Ingreso envase    <valor>
    RUTAS             <valor>
    Merma             <valor>
    Fin SAP           <valor>
    Diferencia        <valor>

El parser solo extrae estos valores. Las filas por SKU en la parte superior
y la tabla auxiliar "INGRESO ENVASE T2" están fuera de scope (se podrán
agregar más adelante).

Salida:
    {
        "t1": {"fisico": <num|None>, "fin_sap": <num|None>, "diferencia": <num|None>},
        "t2": {"inicio_sap": ..., "ingreso_envase": ..., "rutas": ...,
               "merma": ..., "fin_sap": ..., "diferencia": ...},
    }
Si una celda viene vacía, el valor es None.
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl


# Etiqueta esperada en el archivo  →  llave de salida del parser.
T1_FIELDS = {
    "fisico": "Fisico",
    "fin_sap": "Fin SAP",
    "diferencia": "Diferencia",
}

T2_FIELDS = {
    "inicio_sap": "Inicio SAP",
    "ingreso_envase": "Ingreso envase",
    "rutas": "RUTAS",
    "merma": "Merma",
    "fin_sap": "Fin SAP",
    "diferencia": "Diferencia",
}


def _norm(v) -> str:
    """Normaliza para matching: minúsculas, espacios colapsados."""
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def _find_date_tab(sheet_names: list[str], fecha: date) -> Optional[str]:
    """Busca la pestaña con nombre dd.mm.yyyy. Tolera espacios extra."""
    target = fecha.strftime("%d.%m.%Y")
    for name in sheet_names:
        if name.strip() == target:
            return name
    return None


def _find_header_cell(rows, label: str) -> Optional[tuple[int, int]]:
    """Devuelve (row_idx, col_idx) de la primera celda que contenga exactamente `label`."""
    target = _norm(label)
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            if _norm(val) == target:
                return (r_idx, c_idx)
    return None


def _to_number(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s or s.startswith("="):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# Etiquetas que marcan el inicio de OTRO recuadro (no el actual).
# Sirven como "stop" del scan — los recuadros viven adyacentes en la misma
# columna y los labels (Fin SAP, Diferencia) se repiten entre T1 y T2, así
# que sin un terminador el scan de T1 jalaba erróneamente valores de T2.
RECUADRO_HEADERS = {"t1", "t2", "t3"}


def _read_recuadro(rows, header_label: str, fields: dict[str, str]) -> dict[str, Optional[float]]:
    """
    Encuentra la celda que contiene `header_label` (ej. "T1") y luego escanea
    las filas debajo, leyendo etiquetas de la misma columna y valores de la
    columna a la derecha. Se detiene al encontrar otro header de recuadro
    (T1/T2/T3) para no contaminar con valores del recuadro vecino.

    "Primera coincidencia gana" — si un label aparece dos veces, el segundo
    no sobrescribe (defensa adicional).

    Si no encuentra el header, devuelve todos los campos en None.
    """
    out = {key: None for key in fields}
    pos = _find_header_cell(rows, header_label)
    if not pos:
        return out

    r_start, c_label = pos
    c_value = c_label + 1
    label_to_key = {_norm(label): key for key, label in fields.items()}
    self_norm = _norm(header_label)

    for offset in range(1, 15):
        r_idx = r_start + offset
        if r_idx >= len(rows):
            break
        row = rows[r_idx]
        if c_label >= len(row):
            continue
        cell_norm = _norm(row[c_label])
        # Si encontramos OTRO header de recuadro, terminamos el scan.
        if cell_norm in RECUADRO_HEADERS and cell_norm != self_norm:
            break
        if cell_norm in label_to_key:
            key = label_to_key[cell_norm]
            if out[key] is None:  # primera coincidencia gana
                value = row[c_value] if c_value < len(row) else None
                out[key] = _to_number(value)

    return out


def parse_conciliacion(file_path: str | Path, fecha: date) -> dict:
    """
    Lee la pestaña dd.mm.yyyy del archivo y devuelve los dos recuadros.

    Lanza ValueError si la pestaña esperada no existe en el archivo.
    """
    path = Path(file_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = _find_date_tab(wb.sheetnames, fecha)
        if not sheet_name:
            tail = wb.sheetnames[-5:]
            raise ValueError(
                f"No se encontró pestaña '{fecha.strftime('%d.%m.%Y')}' en {path.name}. "
                f"Últimas pestañas: {tail}"
            )
        ws = wb[sheet_name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    return {
        "t1": _read_recuadro(rows, "T1", T1_FIELDS),
        "t2": _read_recuadro(rows, "T2", T2_FIELDS),
    }


if __name__ == "__main__":
    import json
    import sys
    from datetime import datetime

    if len(sys.argv) < 3:
        print("Uso: python parse_conciliacion.py <archivo.xlsx> <YYYY-MM-DD>")
        sys.exit(1)
    f = sys.argv[1]
    fecha = datetime.strptime(sys.argv[2], "%Y-%m-%d").date()
    out = parse_conciliacion(f, fecha)
    print(json.dumps(out, indent=2, ensure_ascii=False, default=str))
