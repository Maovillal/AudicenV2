"""
Tests rápidos para los parsers.

Uso:
    # Con datos sintéticos (no requiere archivos reales):
    python test_parsers.py

    # Con un archivo real:
    python test_parsers.py sap path/al/archivo.XLS [almacen_hint]
    python test_parsers.py conciliacion path/al/archivo.xlsx
    python test_parsers.py classify path/al/archivo
"""

from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

import openpyxl

from classifier import AmbiguousClassificationError, classify
from parse_conciliacion import parse_conciliacion
from parse_sap import parse_sap
from ingest_mail import parse_email_subject


# --- Fixtures sintéticos ---------------------------------------------------

def _make_synthetic_sap_xls(path: Path, almacen: int = 2000) -> None:
    """Genera un .XLS UTF-16-LE TSV imitando una exportación de SAP."""
    headers = ["Ce.", "Material", "Texto breve de material", "Alm.", "UMB",
               "LibrUtiliz", "Bloqueado", "En CtrlCal"]
    body = [
        ["5618", "100123", "CARTA BLANCA 24/355",  str(almacen), "CJ", "1,129.9",   "0",   "0"],
        ["5618", "100124", "TECATE 24/355",        str(almacen), "CJ", "523",       "10",  "0"],
        ["5618", "100125", "INDIO 24/355",         str(almacen), "CJ", "0",         "0",   "5"],
        # Fila de totales (debe filtrarse)
        ["",     "",       "",                     "",           "",   "*",         "",    ""],
    ]
    text = "\t".join(headers) + "\r\n" + "\r\n".join("\t".join(r) for r in body) + "\r\n"
    raw = b"\xff\xfe" + text.encode("utf-16-le")
    path.write_bytes(raw)


def _make_synthetic_conciliacion_xlsx(path: Path) -> None:
    """
    Genera un .xlsx que imita la estructura real del archivo de conciliación:
    - Varias pestañas por fecha (dd.mm.yyyy)
    - Pestañas extra al final ('Conteo 1') que el parser debe IGNORAR
    - Cada pestaña con dos recuadros (T1 y T2) en cols F-G abajo
    """
    wb = openpyxl.Workbook()
    wb.active.title = "07.05.2026"
    wb.active["A1"] = "datos del 7-may"

    # Pestaña del 8-may con los dos recuadros poblados.
    target = wb.create_sheet("08.05.2026")

    # Recuadro T1 (cols F-G, filas 35-38) — como el archivo real.
    target["F35"] = "T1"
    target["F36"] = "Fisico";     target["G36"] = 9514
    target["F37"] = "Fin SAP";    target["G37"] = 9514
    target["F38"] = "Diferencia"; target["G38"] = 0

    # Recuadro T2 — pegado abajo. Sin terminator, T1 jalaría el 'Diferencia' de T2.
    target["F40"] = "T2"
    target["F41"] = "Inicio SAP";     target["G41"] = 9514
    target["F42"] = "Ingreso envase"; target["G42"] = 0
    target["F43"] = "RUTAS";          target["G43"] = None
    target["F44"] = "Merma";          target["G44"] = None
    target["F45"] = "Fin SAP";        target["G45"] = None
    target["F46"] = "Diferencia";     target["G46"] = 9514

    # Pestañas extra al final que el parser DEBE ignorar.
    wb.create_sheet("Hoja2")
    extra = wb.create_sheet("Conteo 1")
    extra["F1"] = "T1"
    extra["F2"] = "Fisico"; extra["G2"] = 99999  # valor trampa

    wb.save(path)


# --- Tests -----------------------------------------------------------------

def test_classifier() -> None:
    cases = [
        ("2000_02_05_2026_INICIO.XLS",          ("sap_2000",            1, "inicio", "2026-05-02")),
        ("2010__TURNO_2_CIERRE_02_05_26.XLS",   ("sap_2010",            2, "cierre", "2026-05-02")),
        ("Inicio_CONCILIACION_ENVASE_2026.xlsx",("conciliacion_envase", 1, "inicio", None)),
        ("CONCILIACION_ENVASE_2026.xlsx",       ("conciliacion_envase", 2, "cierre", None)),
        ("2000_fin_de_turno.XLS",               ("sap_2000",            3, "cierre", None)),
        ("02_2026_TARDE__MAYO__-FIN.xlsx",      ("vertical_liquido",    3, "cierre", "2026-05-02")),
        ("Indicador_nivel_de_Servicio_Rutas_MAYO_2026.xlsx",
                                                ("nivel_servicio",   None, None,     None)),
        ("Ingreso_Envase__MAYO__2026_GP.xlsx",  ("ingreso_envase",   None, None,     None)),
        # Convención nueva: marcador T1/T3 explícito en SAP cierre.
        ("2000_T1_CIERRE_02_05_2026.XLS",       ("sap_2000",            1, "cierre", "2026-05-02")),
        ("2000_T3_CIERRE_02_05_2026.XLS",       ("sap_2000",            3, "cierre", "2026-05-02")),
    ]
    failures = 0
    for filename, expected in cases:
        c = classify(filename)
        actual = (
            c.tipo, c.turno, c.momento,
            c.fecha.isoformat() if c.fecha else None,
        )
        ok = actual == expected
        mark = "OK" if ok else "FAIL"
        print(f"  [{mark}] {filename}")
        if not ok:
            print(f"        esperado: {expected}")
            print(f"        actual:   {actual}")
            failures += 1

    # Caso ambiguo: SAP cierre sin marcador → debe levantar error.
    ambiguous = "2000_CIERRE_02_05_2026.XLS"
    try:
        classify(ambiguous)
    except AmbiguousClassificationError:
        print(f"  [OK] {ambiguous} → AmbiguousClassificationError (esperado)")
    else:
        print(f"  [FAIL] {ambiguous} debió levantar AmbiguousClassificationError")
        failures += 1

    if failures:
        raise AssertionError(f"classifier: {failures} casos fallaron")


def test_parse_sap() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "synthetic_2000.XLS"
        _make_synthetic_sap_xls(path, almacen=2000)
        rows = parse_sap(path, almacen_hint=2000)

    assert len(rows) == 3, f"esperaba 3 filas (sin totales), obtuve {len(rows)}"
    r0 = rows[0]
    assert r0["material"] == "100123"
    assert r0["descripcion"] == "CARTA BLANCA 24/355"
    assert r0["almacen"] == 2000
    assert r0["umb"] == "CJ"
    assert r0["libre_utilizacion"] == 1129.9, f"limpieza de coma falló: {r0}"
    assert r0["bloqueado"] == 0.0
    assert r0["en_ctrl_calidad"] == 0.0
    assert rows[1]["libre_utilizacion"] == 523.0
    assert rows[2]["en_ctrl_calidad"] == 5.0
    print(f"  [OK] parse_sap: {len(rows)} filas, totales filtrados, números limpios")


def test_parse_conciliacion() -> None:
    """Valida que parse_conciliacion encuentra la pestaña por fecha y lee los dos recuadros."""
    from datetime import date as _date

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "synthetic_conciliacion.xlsx"
        _make_synthetic_conciliacion_xlsx(path)
        result = parse_conciliacion(path, _date(2026, 5, 8))

    # T1: ignora la trampa en la pestaña 'Conteo 1'; solo toma de '08.05.2026'.
    assert result["t1"]["fisico"] == 9514, f"T1 fisico: {result}"
    assert result["t1"]["fin_sap"] == 9514, f"T1 fin_sap (no debe contaminarse con T2): {result}"
    assert result["t1"]["diferencia"] == 0, f"T1 diferencia (no debe ser la de T2): {result}"

    # T2: lee independiente sin sobreescribir nada de T1.
    assert result["t2"]["inicio_sap"] == 9514
    assert result["t2"]["ingreso_envase"] == 0
    assert result["t2"]["rutas"] is None
    assert result["t2"]["merma"] is None
    assert result["t2"]["fin_sap"] is None
    assert result["t2"]["diferencia"] == 9514

    # Pestaña inexistente debe levantar error claro.
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "synthetic_conciliacion.xlsx"
        _make_synthetic_conciliacion_xlsx(path)
        try:
            parse_conciliacion(path, _date(2099, 12, 31))
        except ValueError as e:
            assert "31.12.2099" in str(e)
        else:
            raise AssertionError("Debió levantar ValueError por pestaña ausente")

    print("  [OK] parse_conciliacion: pestaña por fecha, dos recuadros sin contaminación")


def test_parse_email_subject() -> None:
    """El asunto del correo es la fuente de verdad para (turno, momento, fecha)."""
    cases = [
        # Convención canónica del supervisor.
        ("INICIO_T1_08.05.2026",       (1, "inicio", "2026-05-08")),
        ("CIERRE_T2_08.05.2026",       (2, "cierre", "2026-05-08")),
        ("CIERRE_T3_31.12.2026",       (3, "cierre", "2026-12-31")),
        # Variaciones de separador.
        ("INICIO T1 08-05-2026",       (1, "inicio", "2026-05-08")),
        ("INICIO-T1-08/05/2026",       (1, "inicio", "2026-05-08")),
        # Año a 2 dígitos.
        ("CIERRE_T2_08.05.26",         (2, "cierre", "2026-05-08")),
        # Sin fecha: turno y momento siguen funcionando.
        ("INICIO_T1",                  (1, "inicio", None)),
        # TURNO_N en lugar de TN.
        ("CIERRE_TURNO_2_08.05.2026",  (2, "cierre", "2026-05-08")),
        # FIN como sinónimo de cierre.
        ("FIN_T3_08.05.2026",          (3, "cierre", "2026-05-08")),
        # Insensible a mayúsculas.
        ("inicio_t1_08.05.2026",       (1, "inicio", "2026-05-08")),
        # Asuntos NO Audicen → todo None (deben filtrarse).
        ("Reunión semanal",            (None, None, None)),
        ("Re: rolas para la fiesta",   (None, None, None)),
        ("",                           (None, None, None)),
    ]
    failures = 0
    for subject, expected in cases:
        t, m, f = parse_email_subject(subject)
        actual = (t, m, f.isoformat() if f else None)
        ok = actual == expected
        mark = "OK" if ok else "FAIL"
        print(f"  [{mark}] {subject!r}")
        if not ok:
            print(f"        esperado: {expected}")
            print(f"        actual:   {actual}")
            failures += 1
    if failures:
        raise AssertionError(f"parse_email_subject: {failures} casos fallaron")


# --- Modo CLI: parsear archivo real ----------------------------------------

def _run_real(mode: str, path: str, *extra: str) -> None:
    if mode == "classify":
        print(json.dumps(classify(path).to_dict(), indent=2, ensure_ascii=False))
    elif mode == "sap":
        hint = int(extra[0]) if extra else None
        out = parse_sap(path, almacen_hint=hint)
        print(f"Filas: {len(out)}")
        print(json.dumps(out[:5], indent=2, ensure_ascii=False, default=str))
    elif mode == "conciliacion":
        out = parse_conciliacion(path)
        print(f"Filas: {len(out)}")
        print(json.dumps(out[:5], indent=2, ensure_ascii=False, default=str))
    else:
        print(f"Modo desconocido: {mode}")
        sys.exit(2)


def main() -> None:
    if len(sys.argv) > 1:
        _run_real(sys.argv[1], sys.argv[2], *sys.argv[3:])
        return

    print("→ parse_email_subject")
    test_parse_email_subject()
    print("→ classifier")
    test_classifier()
    print("→ parse_sap (datos sintéticos)")
    test_parse_sap()
    print("→ parse_conciliacion (datos sintéticos)")
    test_parse_conciliacion()
    print("\nTodos los tests pasaron.")


if __name__ == "__main__":
    main()
